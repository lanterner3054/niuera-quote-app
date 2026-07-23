# -*- coding: utf-8 -*-
"""
NIUERA Customs Pack Engine
--------------------------
报关资料三件套渲染,版式 1:1 对照 DADA 样例
(报关单_DADA.xlsx / 发票_DADA.xlsx / 装箱单_DADA.xlsx;申报要素用户自备)。

generate_customs_pack(config, doc, out_zip_path, base_dir) -> out_zip_path

`doc`(金额/汇总已由调用方算好):
{
  "no": "XNY-AD260015", "date": "2026/03/11",
  "customer": "DaDaCon GmbH", "contact": "Adnan Avci",
  "customer_address": "Westerburger Str. 16, 56459 Winnen, Germany",
  "currency": "USD",
  "shipment": {
    "origin_text": "CHINA", "payment": "T/T", "delivery_terms": "CFR Winnen",
    "qty_unit": "SET", "trade_country": "德国", "arrival_country": "德国",
    "dest_port": "德国", "transport_mode": "航空运输", "trade_mode": "CFR",
    "freight": "521", "insurance": "", "misc_fee": "",
    "package_kind": "纸箱", "marks": "N/M"
  },
  "items": [{"item": "Charging Module", "qty": 2, "unit_price": 775,
             "hs_code": "8504409999", "cn_name": "电源模块", "customs_unit": "个",
             "cartons": 2, "size": "65*47*26", "gw": 43, "nw": 38.4}],
  "stamp": true,
  "totals": {"amount": 1877.0, "cartons": 3, "gw": 53.5, "nw": 47.4}
}

固定报关信息(中文抬头/海关编码/口岸等)一律取 config["customs"],代码不落任何
敏感常量(公开仓库铁律)。
"""
import os
import re
import tempfile
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter, range_boundaries

from engine.pi_engine import _stamp_path, PILImage

# ---------------------------------------------------------------- helpers

THIN = Side(style="thin")
C = Alignment(horizontal="center", vertical="center")
CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left", vertical="center")
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
LT = Alignment(horizontal="left", vertical="top")
LTW = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 发票/装箱单字体(样例为「等线」)
F_CO = Font(name="等线", size=22, bold=True)
F_CO_SUB = Font(name="等线", size=10)
F_DOC_TITLE = Font(name="等线", size=18, bold=True)
F_B = Font(name="等线", size=11, bold=True)
F_N = Font(name="等线", size=11)

# 报关单字体(样例:标签 Noto Sans CJK SC、值 宋体)
F_D_TITLE = Font(name="Noto Sans CJK SC", size=18, bold=True)
F_D_LBL10 = Font(name="Noto Sans CJK SC", size=10)
F_D_LBL = Font(name="Noto Sans CJK SC", size=9)
F_D_VAL = Font(name="宋体", size=9)

STAMP_BOX_W = 190
STAMP_BOX_H = 110


def _safe(s):
    return re.sub(r"[^A-Za-z0-9_.\-一-鿿]", "_", str(s or "").strip())


def _num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def fmt_num(v):
    """775.0 -> '775', 81.75 -> '81.75'(报关单 单价/总价/币制 拼串用)。"""
    f = _num(v)
    if f == int(f):
        return str(int(f))
    return ("%f" % f).rstrip("0").rstrip(".")


_MONTHS = ["January", "February", "March", "April", "May", "June", "July",
           "August", "September", "October", "November", "December"]


def date_en(date_str):
    """'2026/03/11' -> '11th March 2026'(发票/装箱单英文日期)。"""
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", str(date_str or ""))
    if not m:
        return str(date_str or "")
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not 1 <= mo <= 12:
        return str(date_str)
    if 10 <= d % 100 <= 20:
        sfx = "th"
    else:
        sfx = {1: "st", 2: "nd", 3: "rd"}.get(d % 10, "th")
    return f"{d}{sfx} {_MONTHS[mo - 1]} {y}"


def parse_size_cm(size):
    """'65*47*26' / '65x47x26cm' -> (65.0, 47.0, 26.0) or None."""
    nums = re.findall(r"\d+(?:\.\d+)?", str(size or ""))
    if len(nums) != 3:
        return None
    return tuple(float(x) for x in nums)


def compute_customs_totals(items):
    amount = cartons = gw = nw = 0.0
    for it in items:
        amount += _num(it.get("qty")) * _num(it.get("unit_price"))
        cartons += _num(it.get("cartons"))
        gw += _num(it.get("gw"))
        nw += _num(it.get("nw"))
    return {"amount": round(amount, 2), "cartons": int(cartons),
            "gw": round(gw, 2), "nw": round(nw, 2)}


def _box(ws, rng, l=None, r=None, t=None, b=None):
    """给区域内每个格按「区域外框」语义补边框(不清掉已有边)。"""
    c1, r1, c2, r2 = range_boundaries(rng)
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row, col)
            old = cell.border
            cell.border = Border(
                left=THIN if (l and col == c1) else old.left,
                right=THIN if (r and col == c2) else old.right,
                top=THIN if (t and row == r1) else old.top,
                bottom=THIN if (b and row == r2) else old.bottom,
            )


def _grid(ws, rng):
    """区域内全部格子四边细框(明细表用)。"""
    c1, r1, c2, r2 = range_boundaries(rng)
    bd = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row, col).border = bd


def _put(ws, coord, value, font, align=None, merge=None):
    if merge:
        ws.merge_cells(merge)
    cell = ws[coord]
    cell.value = value
    cell.font = font
    if align is not None:
        cell.alignment = align
    return cell


def _insert_stamp(ws, anchor, path):
    try:
        img = XLImage(path)
        w, h = img.width, img.height
        if PILImage is not None:
            try:
                with PILImage.open(path) as im:
                    w, h = im.size
            except Exception:
                pass
        if not w or not h:
            w, h = STAMP_BOX_W, STAMP_BOX_H
        scale = min(STAMP_BOX_W / w, STAMP_BOX_H / h, 1.0)
        img.width = int(w * scale)
        img.height = int(h * scale)
        ws.add_image(img, anchor)
    except Exception:
        pass


def _company_header(ws, config, last_col, title):
    """发票/装箱单共用公司抬头:1 公司名 2 地址 3 联系 5 文件标题。"""
    comp = config.get("company", {})
    cl = get_column_letter(last_col)
    _put(ws, "A1", comp.get("name", ""), F_CO, C, f"A1:{cl}1")
    _put(ws, "A2", comp.get("address", ""), F_CO_SUB, C, f"A2:{cl}2")
    _put(ws, "A3", comp.get("contact_line", ""), F_CO_SUB, C, f"A3:{cl}3")
    ws.merge_cells(f"A4:{cl}4")
    _put(ws, "A5", title, F_DOC_TITLE, C,
         f"A5:{get_column_letter(last_col - 1)}5")
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 14.25


# ---------------------------------------------------------------- 发票

def render_invoice(config, doc, out_path, base_dir):
    """Commercial Invoice,对照 发票_DADA.xlsx:A..L 12 列。"""
    sh = doc.get("shipment", {})
    items = doc.get("items", [])
    cur = doc.get("currency", "USD")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, w in {"A": 11.38, "D": 9.0, "E": 14.25, "G": 8.5, "I": 11.12}.items():
        ws.column_dimensions[col].width = w

    _company_header(ws, config, 12, "Commercial  Invoice")

    ws.merge_cells("A8:D8")
    _put(ws, "A9", "TO:", F_B)
    _put(ws, "A10", doc.get("customer", ""), F_B, L, "A10:E10")
    _put(ws, "G10", f"ORIGIN:       {sh.get('origin_text', '')}", F_B, L, "G10:I10")
    _put(ws, "A11", "\n" + str(doc.get("customer_address", "")), F_B, LTW, "A11:D13")
    _put(ws, "G11", f"PAYMENT:   {sh.get('payment', '')}", F_B, L, "G11:I11")
    _put(ws, "G12", f"TERMS OF DELIVERY:  {sh.get('delivery_terms', '')}", F_B, L, "G12:I12")
    _put(ws, "G13", f"Date:  {date_en(doc.get('date', ''))}", F_B, L, "G13:I13")
    _put(ws, "A14", doc.get("contact", ""), F_B, L, "A14:D14")
    _put(ws, "G14", f"INVOICE NO.:  {doc.get('no', '')}", F_B, L, "G14:I14")
    for rr, h in ((10, 13.5), (11, 19.5), (12, 21.75), (13, 24), (14, 24), (15, 18)):
        ws.row_dimensions[rr].height = h
    for col in ("G", "H", "I"):  # 编号下划线
        ws[f"{col}15"].border = Border(bottom=THIN)

    # ---- 明细表:16 表头,17.. 明细 ----
    HR = 16
    _put(ws, f"A{HR}", "Marks", F_B, C)
    _put(ws, f"B{HR}", "Description of the Goods", F_B, C, f"B{HR}:D{HR}")
    _put(ws, f"E{HR}", f"Quantity ({sh.get('qty_unit', 'SET')})", F_B, C)
    _put(ws, f"F{HR}", f"Unit Price ({cur})", F_B, C, f"F{HR}:G{HR}")
    _put(ws, f"H{HR}", f"Amount ({cur})", F_B, C, f"H{HR}:I{HR}")
    ws.row_dimensions[HR].height = 21

    r = HR + 1
    for it in items:
        _put(ws, f"A{r}", sh.get("marks", "N/M"), F_N, C)
        _put(ws, f"B{r}", it.get("item", ""), F_N, LW, f"B{r}:D{r}")
        _put(ws, f"E{r}", _num(it.get("qty")), F_N, C).number_format = "0"
        c = _put(ws, f"F{r}", _num(it.get("unit_price")), F_N, C, f"F{r}:G{r}")
        c.number_format = "0.00_ "
        c = _put(ws, f"H{r}", f"=F{r}*E{r}", F_B, C, f"H{r}:I{r}")
        c.number_format = "0.00_ "
        ws.row_dimensions[r].height = 58.5
        r += 1
    tr = r
    _put(ws, f"A{tr}", "TOTAL", F_B, C, f"A{tr}:G{tr}")
    c = _put(ws, f"H{tr}", f"=SUM(H{HR + 1}:H{tr - 1})", F_B, C, f"H{tr}:I{tr}")
    c.number_format = "0.00_ "
    _grid(ws, f"A{HR}:I{tr}")

    if doc.get("stamp"):
        sp = _stamp_path(config, base_dir)
        if sp:
            _insert_stamp(ws, f"E{tr + 1}", sp)

    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:L{tr + 8}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- 装箱单

def render_packing_list(config, doc, out_path, base_dir):
    """Packing List,对照 装箱单_DADA.xlsx:A..M 13 列。"""
    sh = doc.get("shipment", {})
    items = doc.get("items", [])
    totals = doc.get("totals", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, w in {"D": 8.88, "E": 10.0, "G": 10.88, "J": 11.25}.items():
        ws.column_dimensions[col].width = w

    _company_header(ws, config, 13, "PACKING LIST")

    _put(ws, "A8", "TO:", F_B, L, "A8:D8")
    _put(ws, "A9", doc.get("customer", ""), F_B, LW, "A9:E9")
    _put(ws, "H9", f"ORIGIN:         {sh.get('origin_text', '')}", F_B, L, "H9:J9")
    _put(ws, "A10", "\n" + str(doc.get("customer_address", "")), F_B, LTW, "A10:D12")
    _put(ws, "H10", f"PAYMENT:     {sh.get('payment', '')}", F_B, L, "H10:J10")
    _put(ws, "H11", f"TERMS OF DELIVERY: {sh.get('delivery_terms', '')}", F_B, L, "H11:J11")
    _put(ws, "H12", f"Date:{date_en(doc.get('date', ''))}", F_B, L, "H12:J12")
    _put(ws, "A13", doc.get("contact", ""), F_B, L, "A13:D13")
    _put(ws, "H13", f"INVOICE NO.:  {doc.get('no', '')}", F_B, L, "H13:J13")
    for rr, h in ((9, 14.85), (10, 13.5), (11, 19.5), (12, 20.25), (13, 17.25),
                  (14, 17.25), (15, 19.5)):
        ws.row_dimensions[rr].height = h
    for col in ("H", "I", "J"):
        ws[f"{col}14"].border = Border(bottom=THIN)

    HR = 15
    _put(ws, f"A{HR}", "Marks", F_B, C)
    _put(ws, f"B{HR}", "Description of the goods", F_B, C, f"B{HR}:D{HR}")
    _put(ws, f"E{HR}", "Quantity", F_B, C)
    _put(ws, f"F{HR}", "Package", F_B, C, f"F{HR}:G{HR}")
    _put(ws, f"H{HR}", "GW (kg)", F_B, C)
    _put(ws, f"I{HR}", "NW (kg)", F_B, C)
    _put(ws, f"J{HR}", "Meas (m³)", F_B, C)

    r = HR + 1
    for it in items:
        cartons = int(_num(it.get("cartons")))
        unit_word = "CARTON" if cartons == 1 else "CARTONS"
        size = str(it.get("size") or "").strip()
        pkg = f"In {cartons} {unit_word}"
        if size:
            pkg += f"                  Size: {size}cm"
        _put(ws, f"A{r}", sh.get("marks", "N/M"), F_N, C)
        _put(ws, f"B{r}", it.get("item", ""), F_N, LW, f"B{r}:D{r}")
        _put(ws, f"E{r}", f"{fmt_num(it.get('qty'))} {sh.get('qty_unit', 'SET')}", F_N, L)
        _put(ws, f"F{r}", pkg, F_N, LW, f"F{r}:G{r}")
        _put(ws, f"H{r}", _num(it.get("gw")), F_N, C)
        _put(ws, f"I{r}", _num(it.get("nw")), F_N, C)
        dims = parse_size_cm(size)
        if dims and cartons:
            meas = "=" + "*".join(fmt_num(d / 100.0) for d in dims)
            if cartons != 1:
                meas += f"*{cartons}"
        else:
            meas = None
        _put(ws, f"J{r}", meas, F_N, C)
        ws.row_dimensions[r].height = 58.5
        r += 1
    tr = r
    cart_total = int(_num(totals.get("cartons")))
    _put(ws, f"A{tr}", "TOTAL", F_B, L)
    _put(ws, f"B{tr}", f"    PACKAGED IN {cart_total} CARTONS", F_B, C, f"B{tr}:G{tr}")
    for col in ("H", "I", "J"):
        _put(ws, f"{col}{tr}", f"=SUM({col}{HR + 1}:{col}{tr - 1})", F_B, C)
    _grid(ws, f"A{HR}:J{tr}")

    if doc.get("stamp"):
        sp = _stamp_path(config, base_dir)
        if sp:
            _insert_stamp(ws, f"F{tr + 1}", sp)

    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:M{tr + 8}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- 报关单

DECL_COL_W = [3.25, 14.5, 6.0, 11.88, 9.62, 3.25, 5.88, 3.38, 5.25, 10.75,
              12.12, 2.75, 5.12, 5.5, 8.88]


def render_declaration(config, doc, out_path):
    """中华人民共和国海关出口货物报关单,对照 报关单_DADA.xlsx:A..O 15 列。
    明细行数可变,行 17..16+n,页脚随之下移。"""
    cus = config.get("customs", {})
    sh = doc.get("shipment", {})
    items = doc.get("items", [])
    totals = doc.get("totals", {})
    cur = doc.get("currency", "USD")

    wb = Workbook()
    ws = wb.active
    ws.title = "出口"
    for j, w in enumerate(DECL_COL_W, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for rr, h in {1: 26.25, 2: 26.25, 3: 28.5, 4: 18, 5: 15, 6: 18, 7: 18.75,
                  8: 14.25, 9: 16.5, 10: 15.75, 11: 14.25, 12: 14.25, 13: 14.25,
                  14: 25.5, 15: 29.25, 16: 13.5}.items():
        ws.row_dimensions[rr].height = h

    LBL, VAL = F_D_LBL, F_D_VAL
    LT_ = Alignment(vertical="top")
    _put(ws, "A1", "中华人民共和国海关出口货物报关单  ", F_D_TITLE, C, "A1:N1")
    _put(ws, "A3", "预录入编号:", F_D_LBL10, Alignment(vertical="center"), "A3:D3")
    _put(ws, "E3", "海关编号:", F_D_LBL10, Alignment(vertical="center"), "E3:G3")
    _put(ws, "I3", "   海关", F_D_LBL10, Alignment(vertical="center"), "I3:L3")
    _put(ws, "M3", "页码/页数", F_D_LBL10, CW, "M3:O3")
    _box(ws, "M3:O3", b=True)

    code = cus.get("customs_code", "")
    shipper = cus.get("shipper_cn", "")
    # 行4-5 境内收发货人 / 出境关别 / 出口日期 / 申报日期 / 备案号
    _put(ws, "A4", f"境内收发货人 {code}".rstrip(), LBL, LT_, "A4:D4")
    _put(ws, "E4", "出境关别", LBL, LT_, "E4:G4")
    _put(ws, "H4", "出口日期", LBL, LT_, "H4:J4")
    _put(ws, "K4", "申报日期", LBL, LT_, "K4:L4")
    _put(ws, "M4", "备案号", LBL, LT_, "M4:O4")
    _put(ws, "A5", shipper, LBL, L, "A5:D5")
    _put(ws, "E5", sh.get("exit_customs", ""), VAL, L, "E5:G5")
    ws.merge_cells("H5:J5"); ws.merge_cells("K5:L5"); ws.merge_cells("M5:O5")
    # 行6-7 境外收货人 / 运输方式 / 运输工具 / 提运单号
    _put(ws, "A6", "境外收货人", LBL, LT)
    _put(ws, "E6", "运输方式", LBL, LT, "E6:G6")
    _put(ws, "H6", "运输工具名称及航次号", LBL, LT, "H6:J6")
    _put(ws, "K6", "提运单号", LBL, LT, "K6:O6")
    _put(ws, "A7", doc.get("customer", ""), F_D_VAL, LT, "A7:D7")
    _put(ws, "E7", sh.get("transport_mode", ""), LBL, LT, "E7:G7")
    ws.merge_cells("H7:J7"); ws.merge_cells("K7:O7")
    # 行8-9 生产销售单位 / 监管方式 / 征免性质 / 许可证号
    _put(ws, "A8", f"生产销售单位     {code}".rstrip(), LBL, LT, "A8:D8")
    _put(ws, "E8", "监管方式", LBL, LT, "E8:G8")
    _put(ws, "H8", "征免性质", LBL, LT, "H8:J8")
    _put(ws, "K8", "许可证号", LBL, LT, "K8:O8")
    _put(ws, "A9", shipper, LBL, LT, "A9:D9")
    _put(ws, "E9", cus.get("supervision_mode", ""), LBL, LT, "E9:G9")
    _put(ws, "H9", cus.get("tax_nature", ""), LBL, LT, "H9:J9")
    ws.merge_cells("K9:O9")
    # 行10-11 合同协议号 / 贸易国 / 运抵国 / 指运港 / 离境口岸
    _put(ws, "A10", "合同协议号", LBL, LT, "A10:D10")
    _put(ws, "E10", "贸易国(地区)", LBL, LT, "E10:G10")
    _put(ws, "H10", "运抵国(地区)", LBL, LT, "H10:J10")
    _put(ws, "K10", "指运港", LBL, LT, "K10:L10")
    _put(ws, "M10", "离境口岸", LBL, LT, "M10:O10")
    _put(ws, "A11", doc.get("no", ""), VAL, LT, "A11:D11")
    _put(ws, "E11", sh.get("trade_country", ""), LBL, Alignment(horizontal="center", vertical="top"), "E11:G11")
    _put(ws, "H11", sh.get("arrival_country", ""), LBL, Alignment(horizontal="center", vertical="top"), "H11:J11")
    _put(ws, "K11", sh.get("dest_port", ""), LBL, Alignment(horizontal="center", vertical="top"), "K11:L11")
    _put(ws, "M11", cus.get("departure_port", ""), LBL, Alignment(horizontal="center", vertical="top"), "M11:O11")
    # 行12-13 包装种类 / 件数 / 毛重 / 净重 / 成交方式 / 运费 / 保费 / 杂费
    _put(ws, "A12", "包装种类", LBL, LT, "A12:D12")
    _put(ws, "E12", "件数", LBL, LT)
    _put(ws, "F12", "毛重(千克)", LBL, LT, "F12:G12")
    _put(ws, "H12", "净重(千克)", LBL, LT)
    _put(ws, "J12", "成交方式", LBL, LT)
    _put(ws, "K12", "运费", LBL, LT)
    _put(ws, "L12", "保费", LBL, LT, "L12:M12")
    _put(ws, "N12", "杂费", LBL, LT, "N12:O12")
    CT = Alignment(horizontal="center", vertical="top")
    freight = str(sh.get("freight") or "").strip()
    _put(ws, "A13", sh.get("package_kind", ""), LBL, LT, "A13:D13")
    _put(ws, "E13", int(_num(totals.get("cartons"))), VAL, CT)
    _put(ws, "F13", _num(totals.get("gw")), VAL, CT, "F13:G13")
    _put(ws, "H13", _num(totals.get("nw")), VAL, CT, "H13:I13")
    _put(ws, "J13", sh.get("trade_mode", ""), VAL, CT)
    _put(ws, "K13", f"{fmt_num(freight)}/{cur}" if freight else "", VAL, CT)
    _put(ws, "L13", str(sh.get("insurance") or ""), VAL, CT, "L13:M13")
    _put(ws, "N13", str(sh.get("misc_fee") or ""), VAL, CT, "N13:O13")
    # 行14 随附单证 / 行15 标记唛码及备注
    _put(ws, "A14", "随附单证", LBL, LTW, "A14:O14")
    marks = str(sh.get("marks") or "").strip()
    _put(ws, "A15", "标记唛码及备注" + ("    " + marks if marks else ""), LBL,
         LTW, "A15:O15")

    # 外框(表头区)
    _box(ws, "A4:O9", l=True, r=True, t=True, b=True)
    for rng in ("A4:D5", "A6:D7", "A8:D9", "E4:G5", "E6:G7", "E8:G9",
                "H4:J5", "H6:J7", "H8:J9", "K4:L5", "M4:O5",
                "K6:O7", "K8:O9"):
        _box(ws, rng, l=True, r=True, t=True, b=True)
    for rng in ("A10:D11", "E10:G11", "H10:J11", "K10:L11", "M10:O11",
                "A12:D13", "E12:E13", "F12:G13", "H12:I13", "J12:J13",
                "K12:K13", "L12:M13", "N12:O13"):
        _box(ws, rng, l=True, r=True, t=True, b=True)
    _box(ws, "A14:O14", l=True, r=True, b=True)
    _box(ws, "A15:O15", l=True, r=True, t=True, b=True)

    # ---- 商品明细:16 表头,17..16+n 明细 ----
    HR = 16
    _put(ws, f"A{HR}", "项号", LBL, C)
    _put(ws, f"B{HR}", "商品编号", LBL, C)
    _put(ws, f"C{HR}", " 商品名称、规格型号  ", LBL, CW, f"C{HR}:D{HR}")
    _put(ws, f"E{HR}", "数量及单位", LBL, C)
    _put(ws, f"F{HR}", "单价/总价/币制", LBL, C, f"F{HR}:I{HR}")
    _put(ws, f"J{HR}", "原产国(地区)", LBL, C)
    _put(ws, f"K{HR}", "最终目的国", LBL, C, f"K{HR}:L{HR}")
    _put(ws, f"M{HR}", "境内货源地", LBL, C, f"M{HR}:N{HR}")
    _put(ws, f"O{HR}", "征免", LBL, C)

    r = HR + 1
    for i, it in enumerate(items, 1):
        qty = _num(it.get("qty"))
        amt = round(qty * _num(it.get("unit_price")), 2)
        _put(ws, f"A{r}", i, VAL, C).number_format = "0"
        _put(ws, f"B{r}", str(it.get("hs_code") or ""), VAL, C)
        _put(ws, f"C{r}", it.get("cn_name", ""), F_D_LBL, CW, f"C{r}:D{r}")
        _put(ws, f"E{r}", f"{fmt_num(qty)}{it.get('customs_unit', '个')}", VAL, C)
        _put(ws, f"F{r}",
             f"{fmt_num(it.get('unit_price'))}/{fmt_num(amt)}/{cur}",
             VAL, C, f"F{r}:I{r}")
        _put(ws, f"J{r}", cus.get("origin_country", "中国"), LBL, C)
        _put(ws, f"K{r}", sh.get("dest_country", sh.get("arrival_country", "")),
             LBL, C, f"K{r}:L{r}")
        _put(ws, f"M{r}", cus.get("domestic_source", ""), LBL, C, f"M{r}:N{r}")
        _put(ws, f"O{r}", cus.get("tax_mode", ""), LBL, C)
        ws.row_dimensions[r].height = 24.75
        r += 1
    _grid(ws, f"A{HR + 1}:O{r - 1}")
    _box(ws, f"A{HR}:O{r - 1}", l=True, r=True, b=True)
    for col_pair in (("A", "A"), ("B", "B"), ("C", "D"), ("E", "E"),
                     ("F", "I"), ("J", "J"), ("K", "L"), ("M", "N"), ("O", "O")):
        _box(ws, f"{col_pair[0]}{HR}:{col_pair[1]}{r - 1}", l=True, r=True)

    # ---- 页脚:报关员 / 声明 / 海关批注 ----
    f1, f2 = r, r + 1
    ws.row_dimensions[f1].height = 24
    ws.row_dimensions[f2].height = 24
    VC = Alignment(vertical="center")
    _put(ws, f"A{f1}", "报关员", LBL, VC, f"A{f1}:B{f1}")
    _put(ws, f"C{f1}", "报关人员证号", LBL, VC, f"C{f1}:D{f1}")
    _put(ws, f"E{f1}", "电话", LBL, VC)
    _put(ws, f"F{f1}", "兹声明以上内容承担如实申报、依法纳税之法律责任", LBL, VC,
         f"F{f1}:K{f1}")
    _put(ws, f"L{f1}", "海关批注及签章", LBL,
         Alignment(horizontal="center", vertical="top"), f"L{f1}:O{f2}")
    ws.merge_cells(f"C{f2}:D{f2}")
    _put(ws, f"F{f2}", "                                申报单位(签章)", LBL, VC,
         f"F{f2}:I{f2}")
    _box(ws, f"A{f1}:K{f2}", l=True, b=True)
    _box(ws, f"L{f1}:O{f2}", l=True, r=True, t=True, b=True)

    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:O{f2}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.39
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.87
    ws.page_margins.bottom = 0.0
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- 打包

def generate_customs_pack(config, doc, out_zip_path, base_dir=None):
    """生成三件套(报关单/发票/装箱单)并打包 zip。返回 out_zip_path。

    自检:doc["totals"]["amount"] 与逐行重算不一致时拒绝出文件。"""
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(out_zip_path))
    items = doc.get("items", [])
    if not items:
        raise ValueError("报关资料不能没有商品行")
    calc = compute_customs_totals(items)
    declared = (doc.get("totals") or {}).get("amount")
    if declared is not None and abs(_num(declared) - calc["amount"]) > 0.005:
        raise ValueError(f"报关金额不一致: 声明 {declared} != 实算 {calc['amount']}")
    doc = dict(doc)
    doc["totals"] = calc

    tag = _safe(doc.get("no", "NA"))
    with tempfile.TemporaryDirectory() as td:
        parts = [
            (f"报关单_{tag}.xlsx",
             render_declaration(config, doc, os.path.join(td, "decl.xlsx"))),
            (f"发票_{tag}.xlsx",
             render_invoice(config, doc, os.path.join(td, "inv.xlsx"), base_dir)),
            (f"装箱单_{tag}.xlsx",
             render_packing_list(config, doc, os.path.join(td, "pack.xlsx"), base_dir)),
        ]
        with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for arcname, path in parts:
                zf.write(path, arcname)
    return out_zip_path
