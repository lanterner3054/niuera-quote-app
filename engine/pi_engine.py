"""
NIUERA PI Engine
----------------
Proforma Invoice renderer, layout 1:1 against the confirmed sample
pi__20260605_XNY-AD260045.pdf. Reuses quote_engine styles/helpers;
the quotation rendering path is untouched.

generate_pi(config, doc, out_path, base_dir) -> out_path

`doc` (fully resolved, amounts already validated by the caller):
{
  "no": "XNY-AD260045", "date": "2026/06/05",
  "customer": "...", "contact": "...", "from_name": "Adrian Ding",
  "currency": "USD",
  "items": [
    {"line_type": "product", "item": "...", "desc": "...", "unit": "PCS",
     "qty": 6, "unit_price": 108.0},
    {"line_type": "fee", "item": "FOB Inland Fee", "desc": "...", "amount": 250.0}
  ],
  "terms": {"origin":..., "payment":..., "price_term":..., "delivery_time":...,
            "package":..., "validity":...},
  "bank_account": {"ac_bank":..., "address":..., "beneficiary":...,
                   "account_no":..., "swift":...},
  "stamp": true,
  "totals": {"amount": 898.0, "amount_words": "EIGHT HUNDRED NINETY EIGHT ONLY"}
}
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from engine.quote_engine import (
    F_TITLE, F_ADDR, F_QUOTE, F_LBL, F_NORM, F_VAL, F_HDR, F_ITEM, F_DESC,
    HDR_FILL, BORD, C, CL, TL, CR, _logo_path, _est_row_height, PILImage,
)
from engine.amount_words import amount_to_words

F_RED = Font(name="Calibri", size=11, color="FF0000")
F_BANK = Font(name="Calibri", size=11, bold=True)
F_WORDS = Font(name="Calibri", size=11, bold=True)

# Grid: A=No. B=Item C:D=Descriptions(merged) E=Unit F=Qty G=Unit Price H=Amount
COL_WIDTHS = [5.5, 24, 26, 26, 8, 8, 16, 15]
NUM_FMT = "0.0##"   # sample shows 6.0 / 108.0 / 648.0 / 898.0

STAMP_BOX_W = 190   # px, matches the blue company-name stamp in the sample
STAMP_BOX_H = 110


def compute_total(items):
    """Authoritative total: product rows qty*unit_price + fee rows amount."""
    total = 0.0
    for it in items:
        if it.get("line_type") == "fee":
            total += float(it.get("amount") or 0)
        else:
            total += float(it.get("qty") or 0) * float(it.get("unit_price") or 0)
    return round(total, 2)


def _write_pi_header(ws, config, doc, base_dir):
    """PI 专属表头:公司名/地址/联系行合并整页宽(A:H)居中,与样例一致。
    (报价单沿用 quote_engine._write_header,互不影响)"""
    comp = config.get("company", {})
    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws["A1"] = comp.get("name", "");         ws["A1"].font = F_TITLE; ws["A1"].alignment = C
    ws["A2"] = comp.get("address", "");      ws["A2"].font = F_ADDR;  ws["A2"].alignment = C
    ws["A3"] = comp.get("contact_line", ""); ws["A3"].font = F_ADDR;  ws["A3"].alignment = C
    ws.row_dimensions[1].height = 70; ws.row_dimensions[2].height = 24; ws.row_dimensions[3].height = 16
    lp = _logo_path(config, base_dir)
    if lp:
        try:
            img = XLImage(lp); img.width = 150; img.height = 88; ws.add_image(img, "A1")
        except Exception:
            pass
    ws.merge_cells("A4:H4"); ws["A4"] = "Proforma Invoice"
    ws["A4"].font = F_QUOTE; ws["A4"].alignment = C
    ws.row_dimensions[4].height = 24
    # To / From block
    ws["A6"] = "To:";   ws["A6"].font = F_LBL
    ws.merge_cells("B6:D6"); ws["B6"] = doc.get("customer", ""); ws["B6"].font = F_VAL; ws["B6"].alignment = CL
    ws["A7"] = "Attn:"; ws["A7"].font = F_LBL
    ws.merge_cells("B7:D7"); ws["B7"] = doc.get("contact", ""); ws["B7"].font = F_VAL; ws["B7"].alignment = CL
    ws["F6"] = "From:"; ws["F6"].font = F_LBL
    ws.merge_cells("G6:H6"); ws["G6"] = doc.get("from_name", ""); ws["G6"].font = F_VAL; ws["G6"].alignment = CL
    ws["F7"] = "Date:"; ws["F7"].font = F_LBL
    ws["G7"] = doc.get("date", ""); ws["G7"].font = F_VAL; ws["G7"].alignment = CL
    ws["F8"] = "NO.:";  ws["F8"].font = F_LBL
    ws.merge_cells("G8:H8"); ws["G8"] = doc.get("no", ""); ws["G8"].font = F_VAL; ws["G8"].alignment = CL
    for rr in (6, 7, 8):
        ws.row_dimensions[rr].height = 18


def _stamp_path(config, base_dir):
    p = (config.get("stamp") or {}).get("path", "stamp.png")
    cand = p if os.path.isabs(p) else os.path.join(base_dir, p)
    return cand if os.path.exists(cand) else None


def _insert_stamp(ws, anchor, path):
    try:
        img = XLImage(path)
        w, h = img.width, img.height
        if PILImage is not None:
            try:
                with PILImage.open(path) as im:
                    w, h = im.size
            except Exception:
                pass
        if not w or not h:
            w, h = STAMP_BOX_W, STAMP_BOX_H
        scale = min(STAMP_BOX_W / w, STAMP_BOX_H / h, 1.0)
        img.width = int(w * scale)
        img.height = int(h * scale)
        ws.add_image(img, anchor)
    except Exception:
        pass


def generate_pi(config, doc, out_path, base_dir=None):
    """Build a Proforma Invoice .xlsx. Returns out_path.

    Raises ValueError if doc["totals"] disagrees with the recomputed total
    (render-consistency self check, plan §7 rule 6)."""
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(out_path))

    items = doc.get("items", [])
    total = compute_total(items)
    declared = (doc.get("totals") or {}).get("amount")
    if declared is not None and abs(float(declared) - total) > 0.005:
        raise ValueError(f"PI 合计不一致: 声明 {declared} != 实算 {total}")
    words = amount_to_words(total)
    declared_words = (doc.get("totals") or {}).get("amount_words")
    if declared_words and declared_words != words:
        raise ValueError("PI 大写金额与合计不一致,拒绝出文件")

    cur = doc.get("currency", "USD")
    wb = Workbook()
    ws = wb.active
    ws.title = "PI"

    _write_pi_header(ws, config, doc, base_dir)
    for j, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- table head (row 10), 7 visual columns on an 8-column grid ----
    HR = 10
    ws.merge_cells(start_row=HR, start_column=3, end_row=HR, end_column=4)
    heads = {1: "No.", 2: "Item", 3: "Descriptions", 5: "Unit", 6: "Qty",
             7: f"Unit Price ({cur})", 8: f"Amount({cur})"}
    for j in range(1, 9):
        c = ws.cell(HR, j, heads.get(j))
        c.font = F_HDR; c.fill = HDR_FILL; c.alignment = C; c.border = BORD
    ws.row_dimensions[HR].height = 26

    # ---- item rows ----
    first = HR + 1
    r = first
    line_no = 0
    for it in items:
        line_no += 1
        ws.cell(r, 1, line_no).font = F_NORM
        ws.cell(r, 1).alignment = C
        ws.cell(r, 1).number_format = "0"
        ws.cell(r, 2, it.get("item", "")).font = F_ITEM
        ws.cell(r, 2).alignment = C
        if it.get("line_type") == "fee":
            # fee row: description spans Descriptions..Unit Price (C..G), sample style
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            d = ws.cell(r, 3, it.get("desc", ""))
            d.font = F_DESC; d.alignment = CL
            a = ws.cell(r, 8, float(it.get("amount") or 0))
            a.font = F_NORM; a.alignment = C; a.number_format = NUM_FMT
            ws.row_dimensions[r].height = max(30, _est_row_height(it.get("desc", "")) - 16)
        else:
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            d = ws.cell(r, 3, it.get("desc", ""))
            d.font = F_DESC; d.alignment = TL
            ws.cell(r, 5, it.get("unit", "PCS")).font = F_NORM
            ws.cell(r, 5).alignment = C
            qty = float(it.get("qty") or 0)
            up = float(it.get("unit_price") or 0)
            for col, val in ((6, qty), (7, up), (8, round(qty * up, 2))):
                cell = ws.cell(r, col, val)
                cell.font = F_NORM; cell.alignment = C; cell.number_format = NUM_FMT
            ws.row_dimensions[r].height = _est_row_height(it.get("desc", ""))
        for j in range(1, 9):
            ws.cell(r, j).border = BORD
        r += 1

    # ---- total row: words centered across B..G, number in H ----
    tr = r
    ws.cell(tr, 1, "Total").font = F_LBL
    ws.cell(tr, 1).alignment = C
    ws.merge_cells(start_row=tr, start_column=2, end_row=tr, end_column=7)
    wc = ws.cell(tr, 2, words)
    wc.font = F_WORDS; wc.alignment = C
    tv = ws.cell(tr, 8, total)
    tv.font = F_WORDS; tv.alignment = C; tv.number_format = NUM_FMT
    for j in range(1, 9):
        ws.cell(tr, j).border = BORD
    ws.row_dimensions[tr].height = 22

    # ---- commerce terms (validity in red) ----
    t = doc.get("terms", {})
    row = tr + 2
    ws.cell(row, 1, "Commerce Terms:").font = F_LBL
    term_lines = [
        (f"1.Origin: {t.get('origin', '')}", F_NORM),
        (f"2.Payment: {t.get('payment', '')}", F_NORM),
        (f"3.Price Term: {t.get('price_term', '')}", F_NORM),
        (f"4.Delivery time: {t.get('delivery_time', '')}", F_NORM),
        (f"5.Package: {t.get('package', '')}", F_NORM),
        (f"6.Validity: {t.get('validity', '')}", F_RED),
    ]
    for k, (line, font) in enumerate(term_lines):
        rr = row + 1 + k
        ws.cell(rr, 1, line).font = font
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)

    # ---- stamp: floats to the right of the terms block ----
    if doc.get("stamp"):
        sp = _stamp_path(config, base_dir)
        if sp:
            _insert_stamp(ws, f"F{row + 1}", sp)

    # ---- bank information ----
    bank = doc.get("bank_account") or {}
    brow = row + len(term_lines) + 2
    ws.cell(brow, 1, "7. Bank Information:").font = F_NORM
    bank_lines = [
        f"A/C BANK: {bank.get('ac_bank', '')}",
        f"ADDRESS: {bank.get('address', '')}",
        f"BENEFICIARY: {bank.get('beneficiary', '')}",
        f"BNF'S A/C: {bank.get('account_no', '')}",
        f"SWIFT CODE: {bank.get('swift', '')}",
    ]
    for k, line in enumerate(bank_lines):
        rr = brow + 1 + k
        ws.cell(rr, 1, line).font = F_BANK
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    end = brow + len(bank_lines) + 1

    # ---- page setup: exact print area => no blank trailing page ----
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:H{end}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A11"
    wb.save(out_path)
    return out_path
