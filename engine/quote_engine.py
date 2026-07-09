"""
NIUERA Quote Engine
-------------------
Pure, UI-independent quote generator. Reuses the proven bilingual layout
(logo, header, dark table head, multi-line descriptions, commerce terms).

generate_quote(config, quote, out_path) -> out_path

`quote` is a fully-resolved dict from the UI (all values already chosen):
{
  "mode": "standard" | "tiered",
  "customer": "Example Trading Ltd",
  "contact": "John Smith",
  "pi_no": "XNY-AD260041",
  "date": "2026/06/03",          # display string
  "from_name": "Adrian Ding",
  "currency": "USD",
  "items": [
     {"item": "...", "desc": "...", "unit": "PCS",
      "qty": 8, "unit_price": 490.0,                 # standard mode
      "band1": "<100", "price1": 490, "band2": ">500", "price2": 460}  # tiered mode
  ],
  "terms": {"origin":..., "payment":..., "price_term":..., "delivery_time":...,
            "package":..., "validity":...}
}
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

# ---- shared styles ----
F_TITLE = Font(name="Calibri", size=22, bold=True)
F_ADDR  = Font(name="Calibri", size=10)
F_QUOTE = Font(name="Calibri", size=16, bold=True)
F_LBL   = Font(name="Calibri", size=11, bold=True)
F_NORM  = Font(name="Calibri", size=11)
F_VAL   = Font(name="Calibri", size=11)
F_HDR   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_ITEM  = Font(name="Calibri", size=11, bold=True)
F_DESC  = Font(name="Calibri", size=9)
F_SMALL = Font(name="Calibri", size=9, italic=True, color="808080")
F_BAND  = Font(name="Calibri", size=10, bold=True)
HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
TOT_FILL  = PatternFill("solid", fgColor="DDEBF7")
BAND_FILL = PatternFill("solid", fgColor="E2EFDA")
_thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
TL = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
CR = Alignment(horizontal="right",  vertical="center")


def _logo_path(config, base_dir):
    p = config.get("company", {}).get("logo", "")
    if not p:
        return None
    cand = p if os.path.isabs(p) else os.path.join(base_dir, p)
    return cand if os.path.exists(cand) else None


def _write_header(ws, config, quote, ncol, base_dir, title="Quotation"):
    comp = config.get("company", {})
    ws.merge_cells("A1:B2")
    ws.merge_cells("C1:H1"); ws.merge_cells("C2:H2"); ws.merge_cells("C3:H3")
    ws["C1"] = comp.get("name", ""); ws["C1"].font = F_TITLE; ws["C1"].alignment = C
    ws["C2"] = comp.get("address", ""); ws["C2"].font = F_ADDR; ws["C2"].alignment = C
    ws["C3"] = comp.get("contact_line", ""); ws["C3"].font = F_ADDR; ws["C3"].alignment = C
    ws.row_dimensions[1].height = 70; ws.row_dimensions[2].height = 24; ws.row_dimensions[3].height = 16
    lp = _logo_path(config, base_dir)
    if lp:
        try:
            img = XLImage(lp); img.width = 150; img.height = 88; ws.add_image(img, "A1")
        except Exception:
            pass
    ws.merge_cells("A4:H4"); ws["A4"] = title; ws["A4"].font = F_QUOTE; ws["A4"].alignment = C
    ws.row_dimensions[4].height = 24
    # To / From block
    ws["A6"] = "To:";   ws["A6"].font = F_LBL
    ws.merge_cells("B6:D6"); ws["B6"] = quote.get("customer", ""); ws["B6"].font = F_VAL; ws["B6"].alignment = CL
    ws["A7"] = "Attn:"; ws["A7"].font = F_LBL
    ws.merge_cells("B7:D7"); ws["B7"] = quote.get("contact", ""); ws["B7"].font = F_VAL; ws["B7"].alignment = CL
    ws["F6"] = "From:"; ws["F6"].font = F_LBL
    ws.merge_cells("G6:H6"); ws["G6"] = quote.get("from_name", ""); ws["G6"].font = F_VAL; ws["G6"].alignment = CL
    ws["F7"] = "Date:"; ws["F7"].font = F_LBL
    ws["G7"] = quote.get("date", ""); ws["G7"].font = F_VAL; ws["G7"].alignment = CL
    ws["F8"] = "NO.:";  ws["F8"].font = F_LBL
    ws.merge_cells("G8:H8"); ws["G8"] = quote.get("pi_no", ""); ws["G8"].font = F_VAL; ws["G8"].alignment = CL
    for rr in (6, 7, 8):
        ws.row_dimensions[rr].height = 18


def _write_terms(ws, quote, start_row):
    t = quote.get("terms", {})
    ws.cell(start_row, 1, "Commerce Terms:").font = F_LBL
    lines = [
        f"1. Origin: {t.get('origin','')}",
        f"2. Payment: {t.get('payment','')}",
        f"3. Price Term: {t.get('price_term','')}",
        f"4. Delivery Time: {t.get('delivery_time','')}",
        f"5. Package: {t.get('package','')}",
        f"6. Validity: {t.get('validity','')}",
    ]
    for k, line in enumerate(lines):
        ws.cell(start_row + 1 + k, 1, line).font = F_NORM
        ws.merge_cells(start_row=start_row + 1 + k, start_column=1, end_row=start_row + 1 + k, end_column=8)
    return start_row + len(lines) + 1


def _est_row_height(desc):
    nl = (desc or "").count("\n") + 1
    return max(46, min(nl * 12 + 8, 170))


PHOTO_BOX_W = 104   # px, fits the ~16-width Photo column
PHOTO_BOX_H = 104   # px


def _resolve_img(img, base_dir):
    if not img:
        return None
    p = img if os.path.isabs(img) else os.path.join(base_dir, "images", img)
    if os.path.exists(p):
        return p
    # also try as a path already relative to base_dir
    p2 = img if os.path.isabs(img) else os.path.join(base_dir, img)
    return p2 if os.path.exists(p2) else None


def _insert_photo(ws, row, img_path, col=3):
    """Insert an image anchored to the Photo cell, scaled to fit the box.
    Returns the image pixel height (so the caller can size the row), or 0."""
    if not img_path:
        return 0
    try:
        img = XLImage(img_path)
        w, h = img.width, img.height
        if PILImage is not None:
            try:
                with PILImage.open(img_path) as im:
                    w, h = im.size
            except Exception:
                pass
        if not w or not h:
            w, h = PHOTO_BOX_W, PHOTO_BOX_H
        scale = min(PHOTO_BOX_W / w, PHOTO_BOX_H / h, 1.0)
        img.width = int(w * scale)
        img.height = int(h * scale)
        cell = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell)
        return img.height
    except Exception:
        return 0


def _build_standard(ws, config, quote, base_dir):
    cur = quote.get("currency", "USD")
    _write_header(ws, config, quote, 8, base_dir)
    for j, w in enumerate([5.5, 30, 16, 46, 7, 7, 15, 15], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    HR = 10
    heads = ["No.", "Item", "Photo", "Descriptions", "Unit", "Qty", f"Unit Price\n({cur})", f"Amount\n({cur})"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(HR, j, h); c.font = F_HDR; c.fill = HDR_FILL; c.alignment = C; c.border = BORD
    ws.row_dimensions[HR].height = 32
    items = quote.get("items", [])
    first = HR + 1
    for i, it in enumerate(items):
        r = first + i
        ws.cell(r, 1, i + 1).font = F_NORM; ws.cell(r, 1).alignment = C; ws.cell(r, 1).number_format = "0"
        ws.cell(r, 2, it.get("item", "")).font = F_ITEM; ws.cell(r, 2).alignment = CL
        d = ws.cell(r, 4, it.get("desc", "")); d.font = F_DESC; d.alignment = TL
        ws.cell(r, 5, it.get("unit", "PCS")).font = F_NORM; ws.cell(r, 5).alignment = C
        qty = it.get("qty")
        up = it.get("unit_price")
        qc = ws.cell(r, 6, qty if qty not in (None, "") else None); qc.font = F_NORM; qc.alignment = C; qc.number_format = "0"
        gc = ws.cell(r, 7, up if up not in (None, "") else None); gc.font = F_NORM; gc.alignment = C; gc.number_format = "#,##0.00"
        if qty not in (None, "") and up not in (None, ""):
            hc = ws.cell(r, 8, f"=F{r}*G{r}")
        else:
            hc = ws.cell(r, 8, None)
        hc.font = F_NORM; hc.alignment = CR; hc.number_format = "#,##0.00"
        for j in range(1, 9):
            ws.cell(r, j).border = BORD
        img_h = _insert_photo(ws, r, _resolve_img(it.get("image"), base_dir), col=3)
        ws.row_dimensions[r].height = max(_est_row_height(it.get("desc", "")), (img_h * 0.78 + 8) if img_h else 0)
    last = first + len(items) - 1 if items else HR
    # total
    tr = last + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=7)
    ws.cell(tr, 1, f"Total ({cur})").font = F_LBL; ws.cell(tr, 1).alignment = CR
    if items:
        tv = ws.cell(tr, 8, f"=SUM(H{first}:H{last})")
    else:
        tv = ws.cell(tr, 8, None)
    tv.font = Font(name="Calibri", size=11, bold=True); tv.alignment = CR; tv.number_format = "#,##0.00"
    for j in range(1, 9):
        ws.cell(tr, j).border = BORD; ws.cell(tr, j).fill = TOT_FILL
    end = _write_terms(ws, quote, tr + 2)
    return end


def _build_tiered(ws, config, quote, base_dir):
    cur = quote.get("currency", "USD")
    _write_header(ws, config, quote, 9, base_dir)
    for j, w in enumerate([5.5, 28, 15, 40, 7, 12, 13, 12, 13], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    HR = 10
    heads = ["No.", "Item", "Photo", "Descriptions", "Unit",
             "Qty Band \u2460", f"Unit Price \u2460\n({cur})", "Qty Band \u2461", f"Unit Price \u2461\n({cur})"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(HR, j, h); c.font = F_HDR; c.fill = HDR_FILL; c.alignment = C; c.border = BORD
    ws.row_dimensions[HR].height = 34
    items = quote.get("items", [])
    first = HR + 1
    for i, it in enumerate(items):
        r = first + i
        ws.cell(r, 1, i + 1).font = F_NORM; ws.cell(r, 1).alignment = C; ws.cell(r, 1).number_format = "0"
        ws.cell(r, 2, it.get("item", "")).font = F_ITEM; ws.cell(r, 2).alignment = CL
        d = ws.cell(r, 4, it.get("desc", "")); d.font = F_DESC; d.alignment = TL
        ws.cell(r, 5, it.get("unit", "PCS")).font = F_NORM; ws.cell(r, 5).alignment = C
        b1 = ws.cell(r, 6, it.get("band1") or ""); b1.font = F_BAND; b1.alignment = C; b1.fill = BAND_FILL
        p1 = it.get("price1")
        c7 = ws.cell(r, 7, p1 if p1 not in (None, "") else None); c7.font = F_NORM; c7.alignment = C; c7.number_format = "#,##0.00"
        b2 = ws.cell(r, 8, it.get("band2") or ""); b2.font = F_BAND; b2.alignment = C; b2.fill = BAND_FILL
        p2 = it.get("price2")
        c9 = ws.cell(r, 9, p2 if p2 not in (None, "") else None); c9.font = F_NORM; c9.alignment = C; c9.number_format = "#,##0.00"
        for j in range(1, 10):
            ws.cell(r, j).border = BORD
        img_h = _insert_photo(ws, r, _resolve_img(it.get("image"), base_dir), col=3)
        ws.row_dimensions[r].height = max(_est_row_height(it.get("desc", "")), (img_h * 0.78 + 8) if img_h else 0)
    last = first + len(items) - 1 if items else HR
    note = last + 1
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=9)
    ws.cell(note, 1, "* Unit prices are tiered by purchase quantity; final price confirmed per PO. "
                     "单价随采购数量分档,最终单价以订单数量确认。").font = F_SMALL
    end = _write_terms(ws, quote, note + 2)
    return end


def generate_quote(config, quote, out_path, base_dir=None):
    """Build an .xlsx quotation file. Returns out_path."""
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(out_path))
    mode = quote.get("mode", "standard")
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    if mode == "tiered":
        end = _build_tiered(ws, config, quote, base_dir)
        printcols = "I"
    else:
        end = _build_standard(ws, config, quote, base_dir)
        printcols = "H"
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:{printcols}{end}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A11"
    wb.save(out_path)
    return out_path
