# -*- coding: utf-8 -*-
"""报关资料引擎测试:四件套结构按 DADA 样例逐项校验。"""
import os
import sys
import zipfile

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from engine.customs_engine import (  # noqa: E402
    compute_customs_totals, date_en, fmt_num, generate_customs_pack,
    parse_size_cm,
)

CONFIG = {
    "company": {
        "name": "Suzhou Niuera Energy Co.,Ltd",
        "address": "No. 81 Chunlan Road, Suzhou, China",
        "contact_line": "T:+86-512-000  Web: x  E-mail: y",
    },
    "customs": {
        "shipper_cn": "测试公司(苏州)有限公司",
        "customs_code": "TESTCODE01",
        "departure_port": "苏州其他",
        "domestic_source": "苏州其他",
        "origin_country": "中国",
        "supervision_mode": "一般贸易",
        "tax_nature": "一般征税",
        "tax_mode": "照章征税",
    },
    "stamp": {"path": "no_such_stamp.png"},
}


def make_doc(n_items=2):
    items = [
        {"item": "Charging Module", "qty": 2, "unit_price": 775,
         "hs_code": "8504409999", "cn_name": "电源模块", "customs_unit": "个",
         "cartons": 2, "size": "65*47*26", "gw": 43, "nw": 38.4},
        {"item": "Cables", "qty": 4, "unit_price": 81.75,
         "hs_code": "8544422900", "cn_name": "连接线缆", "customs_unit": "个",
         "cartons": 1, "size": "65*47*26", "gw": 10.5, "nw": 9},
        {"item": "Extra Part", "qty": 1, "unit_price": 100,
         "hs_code": "1234567890", "cn_name": "配件", "customs_unit": "件",
         "cartons": 1, "size": "", "gw": 5, "nw": 4},
    ][:n_items]
    return {
        "no": "XNY-AD260015", "doc_type": "customs", "revision": 0,
        "date": "2026/03/11",
        "customer": "DaDaCon GmbH", "contact": "Adnan Avci",
        "customer_address": "Westerburger Str. 16, 56459 Winnen, Germany",
        "currency": "USD",
        "shipment": {
            "origin_text": "CHINA", "payment": "T/T",
            "delivery_terms": "CFR Winnen", "qty_unit": "SET",
            "trade_country": "德国", "arrival_country": "德国",
            "dest_port": "德国", "transport_mode": "航空运输",
            "trade_mode": "CFR", "freight": "521", "insurance": "",
            "misc_fee": "", "package_kind": "纸箱", "marks": "N/M",
            "exit_customs": "",
        },
        "items": items,
        "stamp": False,
        "totals": compute_customs_totals(items),
    }


@pytest.fixture
def pack(tmp_path):
    """生成 zip 并解包,返回 {文件名前缀: 路径}。"""
    zp = tmp_path / "pack.zip"
    generate_customs_pack(CONFIG, make_doc(), str(zp))
    out = {}
    with zipfile.ZipFile(zp) as zf:
        names = zf.namelist()
        zf.extractall(tmp_path / "x")
        for n in names:
            out[n.split("_")[0]] = str(tmp_path / "x" / n)
    out["__names__"] = names
    return out


def test_zip_contains_three_parts(pack):
    names = pack["__names__"]
    assert len(names) == 3
    for prefix in ("报关单", "发票", "装箱单"):
        assert any(n.startswith(prefix) for n in names), prefix
    assert all("XNY-AD260015" in n for n in names)


# ---------------------------------------------------------------- helpers


def test_fmt_num():
    assert fmt_num(775) == "775"
    assert fmt_num(775.0) == "775"
    assert fmt_num(81.75) == "81.75"
    assert fmt_num(1550.0) == "1550"


def test_date_en():
    assert date_en("2026/03/11") == "11th March 2026"
    assert date_en("2026/03/01") == "1st March 2026"
    assert date_en("2026-06-22") == "22nd June 2026"
    assert date_en("2026/12/03") == "3rd December 2026"
    assert date_en("garbage") == "garbage"


def test_parse_size_cm():
    assert parse_size_cm("65*47*26") == (65.0, 47.0, 26.0)
    assert parse_size_cm("65x47x26cm") == (65.0, 47.0, 26.0)
    assert parse_size_cm("") is None
    assert parse_size_cm("65*47") is None


def test_compute_totals():
    t = compute_customs_totals(make_doc()["items"])
    assert t == {"amount": 1877.0, "cartons": 3, "gw": 53.5, "nw": 47.4}


def test_total_mismatch_rejected(tmp_path):
    doc = make_doc()
    doc["totals"] = {"amount": 999.0}
    with pytest.raises(ValueError):
        generate_customs_pack(CONFIG, doc, str(tmp_path / "bad.zip"))


def test_empty_items_rejected(tmp_path):
    doc = make_doc()
    doc["items"] = []
    with pytest.raises(ValueError):
        generate_customs_pack(CONFIG, doc, str(tmp_path / "bad.zip"))


# ---------------------------------------------------------------- 发票


def test_invoice_layout(pack):
    ws = load_workbook(pack["发票"]).active
    assert "Commercial" in ws["A5"].value
    assert ws["A1"].value == CONFIG["company"]["name"]
    assert ws["A10"].value == "DaDaCon GmbH"
    assert "Winnen, Germany" in ws["A11"].value
    assert ws["G10"].value.startswith("ORIGIN:")
    assert "CHINA" in ws["G10"].value
    assert ws["G13"].value == "Date:  11th March 2026"
    assert ws["G14"].value == "INVOICE NO.:  XNY-AD260015"
    # 表头与明细
    assert ws["E16"].value == "Quantity (SET)"
    assert ws["F16"].value == "Unit Price (USD)"
    assert ws["A17"].value == "N/M"
    assert ws["B17"].value == "Charging Module"
    assert ws["E17"].value == 2
    assert ws["F17"].value == 775
    assert ws["H17"].value == "=F17*E17"
    assert ws["H18"].value == "=F18*E18"
    assert ws["A19"].value == "TOTAL"
    assert ws["H19"].value == "=SUM(H17:H18)"
    merges = {str(m) for m in ws.merged_cells.ranges}
    for m in ("B16:D16", "F16:G16", "H16:I16", "A11:D13", "A19:G19", "H19:I19"):
        assert m in merges, m


def test_invoice_row_scaling(tmp_path):
    zp = tmp_path / "p3.zip"
    generate_customs_pack(CONFIG, make_doc(3), str(zp))
    with zipfile.ZipFile(zp) as zf:
        inv = [n for n in zf.namelist() if n.startswith("发票")][0]
        zf.extract(inv, tmp_path)
    ws = load_workbook(str(tmp_path / inv)).active
    assert ws["B19"].value == "Extra Part"
    assert ws["A20"].value == "TOTAL"
    assert ws["H20"].value == "=SUM(H17:H19)"


# ---------------------------------------------------------------- 装箱单


def test_packing_layout(pack):
    ws = load_workbook(pack["装箱单"]).active
    assert "PACKING LIST" in ws["A5"].value
    assert ws["H13"].value == "INVOICE NO.:  XNY-AD260015"
    assert ws["J15"].value == "Meas (m³)"
    assert ws["E16"].value == "2 SET"
    assert ws["F16"].value.startswith("In 2 CARTONS")
    assert "Size: 65*47*26cm" in ws["F16"].value
    assert ws["F17"].value.startswith("In 1 CARTON")
    assert ws["J16"].value == "=0.65*0.47*0.26*2"
    assert ws["J17"].value == "=0.65*0.47*0.26"
    assert ws["A18"].value == "TOTAL"
    assert ws["B18"].value.strip() == "PACKAGED IN 3 CARTONS"
    assert ws["H18"].value == "=SUM(H16:H17)"
    assert ws["I18"].value == "=SUM(I16:I17)"
    assert ws["J18"].value == "=SUM(J16:J17)"


# ---------------------------------------------------------------- 报关单


def test_declaration_layout(pack):
    ws = load_workbook(pack["报关单"]).active
    assert ws.title == "出口"
    assert "报关单" in ws["A1"].value
    assert ws["A4"].value == "境内收发货人 TESTCODE01"
    assert ws["A5"].value == "测试公司(苏州)有限公司"
    assert ws["A7"].value == "DaDaCon GmbH"
    assert ws["E7"].value == "航空运输"
    assert ws["E9"].value == "一般贸易"
    assert ws["H9"].value == "一般征税"
    assert ws["A11"].value == "XNY-AD260015"
    assert ws["E11"].value == "德国"
    assert ws["M11"].value == "苏州其他"
    # 件数/毛重/净重/成交方式/运费
    assert ws["A13"].value == "纸箱"
    assert ws["E13"].value == 3
    assert ws["F13"].value == 53.5
    assert ws["H13"].value == 47.4
    assert ws["J13"].value == "CFR"
    assert ws["K13"].value == "521/USD"
    # 商品行
    assert ws["A17"].value == 1
    assert ws["B17"].value == "8504409999"
    assert ws["C17"].value == "电源模块"
    assert ws["E17"].value == "2个"
    assert ws["F17"].value == "775/1550/USD"
    assert ws["J17"].value == "中国"
    assert ws["K17"].value == "德国"
    assert ws["O17"].value == "照章征税"
    assert ws["B18"].value == "8544422900"
    assert ws["F18"].value == "81.75/327/USD"
    # 页脚(2 行明细 -> 19/20 行)
    assert ws["A19"].value == "报关员"
    assert ws["L19"].value == "海关批注及签章"
    assert "申报单位" in ws["F20"].value


def test_declaration_footer_shifts(tmp_path):
    zp = tmp_path / "p3.zip"
    generate_customs_pack(CONFIG, make_doc(3), str(zp))
    with zipfile.ZipFile(zp) as zf:
        decl = [n for n in zf.namelist() if n.startswith("报关单")][0]
        zf.extract(decl, tmp_path)
    ws = load_workbook(str(tmp_path / decl)).active
    assert ws["A19"].value == 3
    assert ws["E19"].value == "1件"
    assert ws["A20"].value == "报关员"
    assert ws["L20"].value == "海关批注及签章"
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert "L20:O21" in merges
    assert "F20:K20" in merges
